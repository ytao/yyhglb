import openpyxl
import os
import conf as c
import pdb

class project_data():
    def __init__(self):
        '''初始化'''
        self.filename='//10.35.5.9/维修日报/三焊装/三焊装.一元化管理表.xlsx'
        # self.filename='一元化管理表.xlsx'
        if os.path.isfile(self.filename)==False:
            self.init_success=False
        if os.path.isfile(self.filename+'.lock')==True:
            self.init_success=False
        else:
            self.record_sheet_name='故障登记'

            self.record_start_row=5
            self.record_max_row=500

            self.record_start_column=3
            self.record_max_column=51

            self.selectList_sheet_name='DataList'

            self.selectList_start_row=3
            self.selectList_max_row=500

            self.selectList_start_column=1
            self.selectList_max_column=21

            self.person_start_column=34
            self.person_max_column=51
            self.person_start_row=self.record_start_row-1

            self.wb=openpyxl.load_workbook(self.filename)
            self.ws=self.wb[self.record_sheet_name]
            self.ws_DataList=self.wb[self.selectList_sheet_name]

            self.init_success=True

    def testWrite2(self,income_information,income_information2,remote_row):
        '''测试功能'''
        this_row=0
        if (remote_row==-2):
            for x in range(self.record_start_row,self.record_max_row):
                if self.ws.cell(row=x,column=self.record_start_column).value:
                    pass
                else:
                    this_row=x
                    break
            else:
                this_row=remote_row+self.record_start_row
        else:
            this_row=int(remote_row)+self.record_start_row
        for i in income_information.keys():
            self.ws.cell(row=this_row,column=c.dic_position[i]).value=income_information[i]
        for i in income_information2.keys():
            self.ws.cell(row=this_row,column=c.dic_position[i]).value=income_information2[i]
        self.wb.save(filename=self.filename)


    def OLDreadRecords(self,l):
        '''读取所有记录并返回的一个列表中'''
        for y in range(self.record_start_row,self.record_max_row):
            if self.ws.cell(row=y,column=self.record_start_column).value:
                for r in c.dic_value.keys():
                    mstr=self.ws.cell(row=y,column=c.dic_position[r]).value
                    # pdb.set_trace()
                    if mstr==None:
                        mstr=''
                    c.dic_value[r] =mstr
                l.append(c.dic_value.copy())
            else:
                break

    def readOneRecord(self,remote_row):
        '''
        读取一行数据，可能是空数据
        '''
        this_row=int(remote_row)
        if (this_row==-2):
            for y in range(self.record_start_row,self.record_max_row):
                if not self.ws.cell(row=y,column=self.record_start_row).value:
                    my_row=y
                    break
                else:
                    my_row=int(remote_row)+self.record_start_row
        else:
            my_row=int(remote_row)+self.record_start_row

        for r in c.dic_value.keys():
            mstr=self.ws.cell(row=my_row,column=c.dic_position[r]).value
            if mstr==None:
                mstr=''
            else:
                mstr=str(mstr)
            c.dic_value[r] =mstr
        return c.dic_value.copy()

    def readDataList(self):
        '''
        读取配置内容,返回值是一个列表
        '''
        this_max_row=self.selectList_max_row
        confs={}
        cur_lst=[]
        remote_datas=[]

        # 得到表格的标题:直接把第二行的数据取出来作为标题
        for x in range(self.selectList_start_column,self.selectList_max_column+1):
            mstr=''
            mstr=str(self.ws_DataList.cell(row=self.selectList_start_row-1,column=x).value)
            if (mstr=='None'):
                break;
            else:
                mstr_caption=''
                for m in c.dic_caption:
                    if (c.dic_caption[m]==mstr):
                        mstr_caption=m
                        confs[mstr_caption]=''
                        break;
                cur_lst=[]

                # 得到每一列数据的最后一行
                for yy in range(self.selectList_max_row,self.selectList_start_row-1,-1):
                    if (self.ws_DataList.cell(row=yy,column=x).value!=None):
                        this_max_row=yy
                        # print(yy)
                        break;

                # 得到每一行的数据，然后逐行写入到一个列表中
                for y in range(self.selectList_start_row,this_max_row+1):
                    mstr_datalist=str(self.ws_DataList.cell(row=y,column=x).value)
                    if (mstr_datalist=='None'):
                        mstr_datalist=''
                    else:
                        cur_lst.append(mstr_datalist)
                confs[mstr_caption]=cur_lst.copy()

        return confs.copy()


    def getRecords(self):
        '''
        得到所有的数据记录,返回值分别是标题和数据
        '''
        this_max_row=0
        remote_captions=[]
        cur_lst=[]
        remote_datas=[]

        # 得到表格的标题:直接把第二行的数据取出来作为标题
        for x in range(self.record_start_column,self.record_max_column+1):
            if (not self.ws.cell(row=self.record_start_row-1,column=x).value):
                self.record_max_column=x-1;
                self.people_max_column=x-1;
                break;
            else:
                remote_captions.append(self.ws.cell(row=self.record_start_row-1,column=x).value)

        # 得到每一行的数据，然后逐行写入到一个列表中
        for y in range(self.record_start_row,self.record_max_row):
            # 如果那一行已经没有数据了，就直接退出
            if (self.ws.cell(row=y,column=self.record_start_row).value==None):
                break

            # 从第三行到
            cur_lst=[]
            for x in range(self.record_start_column,self.record_max_column+1):
                mstr=str(self.ws.cell(row=y,column=x).value)
                if (mstr=='None'):
                    mstr=''
                cur_lst.append(mstr)
            remote_datas.append(cur_lst.copy())
        # print(len(remote_captions))
        return remote_captions.copy(),remote_datas.copy()

    def getPeopleList(self):
        # 先得到所有人员名称
        list_persons=[]
        for x in range(self.person_start_column,self.record_max_column+1):
            mstr=str(self.ws.cell(row=self.person_start_row,column=x).value)
            if (mstr=='None'):
                self.person_max_column=x-1
                break;
            else:
                list_persons.append(mstr)
        return list_persons.copy()

    def getPeopleStatus(self,remote_row):
        '''
        得到当前行的人员信息
        '''

        dic_persons={}
        list_persons=[]
        this_row=int(remote_row)

        for xx in range(self.person_start_column,self.person_max_column+1):
            my_person_status=str(self.ws.cell(row=this_row+self.record_start_row,column=xx).value)
            if (str(my_person_status)=='1'):
                dic_persons[self.ws.cell(row=self.person_start_row,column=xx).value]='1'
            else:
                dic_persons[self.ws.cell(row=self.person_start_row,column=xx).value]=''

        return dic_persons.copy()

    def writePeopleStatus(self,income_dic,remote_row):
        '''
        写入当前更改的状态
        '''
        this_row=int(remote_row)+self.record_start_row
        this_list=self.getPeopleList()
        for x in range(self.person_start_column,self.person_max_column+1):
            for a in this_list:
                if (a==self.ws.cell(row=self.person_start_row,column=x).value):
                    self.ws.cell(row=this_row,column=x).value=income_dic[a];
                    continue;
